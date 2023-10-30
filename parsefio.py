#!/usr/bin/env python3

import logging
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
import argparse

# Logging Configuration
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    filename='fio_processing.log',
                    filemode='w')

logging.info("Script started.")


def convert_latencies(latency_in_nsecs):
    """Convert latencies from nanoseconds to milliseconds."""
    return round(latency_in_nsecs / 1_000_000, 2)


def convert_bandwidth(bandwidth_in_kib):
    """Convert bandwidth from KiB/sec to MiB/sec."""
    return round(bandwidth_in_kib / 1024, 2)


def write_to_csv(data_df: pd.DataFrame, filename: str):
    """Write the data DataFrame to a CSV file."""
    logging.info(f"Writing data to CSV file: {filename}")
    data_df.to_csv(filename, index=False)
    logging.info(f"Data successfully written to {filename}")
    return None


def write_to_excel(data_df: pd.DataFrame, filename: str):
    """Write the data DataFrame to an Excel file."""
    logging.info(f"Writing data to Excel file: {filename}")
    wb = Workbook()
    ws = wb.active

    # Convert DataFrame to dictionary
    data_dict = data_df.to_dict(orient='list')

    # Format headers
    def format_header(header):
        # If the header is a tuple, join its elements with underscores
        if isinstance(header, tuple):
            return "_".join([str(x) for x in header if x])
        return str(header)

    headers = [format_header(header) for header in data_dict.keys()]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for col_num, key in enumerate(data_dict.keys(), 1):
        for row_num, value in enumerate(data_dict[key], 2):  # Start from row 2 as headers are in row 1
            ws.cell(row=row_num, column=col_num, value=float(value))  # Convert value to string

    wb.save(filename)
    logging.info(f"Data successfully written to {filename}")
    return None


def process_log_file(filename, directory):
    parts = filename.split("_")
    job_name = parts[0]
    job_type = parts[1].split(".")[0]

    df = pd.read_csv(os.path.join(directory, filename), header=None, delimiter=',',
                     names=["Time", "Value", "Data_direction", "Blk_size", "offset", "cmd_pri"])
    df.drop(columns=["Time", "Blk_size", "offset", "cmd_pri"], inplace=True)

    # Strip white spaces
    df = df.apply(lambda col: col.str.strip() if col.dtype == "object" else col)

    # Convert data_direction to readable format
    df["Data_direction"] = df["Data_direction"].map({0: "Read", 1: "Write", 2: "Trim"}).fillna("Unknown")

    # Add job details to the dataframe
    df["Job_name"] = job_name.strip()
    df["Job_type"] = job_type.strip()
    df = df.reset_index(drop=True)
    df["Row_num"] = df.index + 1  # +1 to make it 1-based

    return df, job_type


def process_standalone_log_files(directory):
    logging.info("Starting the processing of log files in directory: %s", directory)
    all_dataframes = []
    unique_job_types = set()  # To keep track of unique job types
    aggregated_dataframes = []
    for filename in sorted(os.listdir(directory)):
        if filename.endswith(".log"):
            logging.info("Processing file: %s", filename)
            df, job_type = process_log_file(filename, directory)
            unique_job_types.add(job_type)
            all_dataframes.append(df)
    logging.info("Aggregating data from all files...")
    master_df = pd.concat(all_dataframes)
    # Convert the Value column to float
    master_df["Value"] = master_df["Value"].astype(float)
    for job_type in unique_job_types:
        logging.debug(f"Job Type :{job_type}\n")
        # Convert values based on job_type before aggregation
        mask_latency = (master_df["Job_type"] == job_type) & (job_type in ["clat", "slat", "lat"])
        mask_bandwidth = (master_df["Job_type"] == job_type) & (job_type == "bw")
        master_df.loc[mask_latency, "Value"] = master_df.loc[mask_latency, "Value"].apply(convert_latencies)
        master_df.loc[mask_bandwidth, "Value"] = master_df.loc[mask_bandwidth, "Value"].apply(convert_bandwidth)
        aggregated_dataframes.append(master_df)
    aggregated_df = pd.concat(aggregated_dataframes)
    result_df = aggregated_df.pivot_table(index="Row_num", columns=["Job_name", "Job_type", "Data_direction"],
                                          values="Value", aggfunc="first")
    result_df = result_df.reset_index()
    logging.debug(f"Missing values in result DataFrame:\n{result_df.isnull().sum()}")
    result_df.fillna(0, inplace=True)
    logging.debug(f"Final Result DataFrame:\n{result_df.head()}")
    logging.info(f"Finished processing client/server log files in directory: {directory}")
    return result_df


def process_client_server_log_files(directory):
    logging.info("Starting processing of log files in client/server mode.")

    # Initialization
    all_dataframes = []
    unique_job_types = set()  # To keep track of unique job types
    aggregated_dataframes = []
    temp_df = None

    for filename in sorted(os.listdir(directory)):
        if ".log." in filename:
            logging.debug(f"Processing data for file: {filename}")
            df, job_type = process_log_file(filename, directory)
            unique_job_types.add(job_type)
            all_dataframes.append(df)

    # Concatenate all dataframes
    master_df = pd.concat(all_dataframes)
    logging.debug(f"Master DataFrame after concatenation:\n{master_df}")

    # Group by Row_num, Job_name, Job_type, and Data_direction and aggregate
    # Convert the Value column to float
    master_df["Value"] = master_df["Value"].astype(float)
    for job_type in unique_job_types:
        logging.debug(f"Job Type :{job_type}\n")

        # Convert values based on job_type before aggregation
        mask_latency = (master_df["Job_type"] == job_type) & (job_type in ["clat", "slat", "lat"])
        mask_bandwidth = (master_df["Job_type"] == job_type) & (job_type == "bw")

        master_df.loc[mask_latency, "Value"] = master_df.loc[mask_latency, "Value"].apply(convert_latencies)
        master_df.loc[mask_bandwidth, "Value"] = master_df.loc[mask_bandwidth, "Value"].apply(convert_bandwidth)

        if job_type in ["clat", "slat", "lat"]:
            temp_df = master_df[master_df["Job_type"] == job_type].groupby(
                ["Row_num", "Job_name", "Job_type", "Data_direction"]).mean().reset_index()
            logging.debug(f"Mean Aggregated DataFrame:\n{temp_df}")
        elif job_type in ["bw", "iops"]:
            temp_df = master_df[master_df["Job_type"] == job_type].groupby(
                ["Row_num", "Job_name", "Job_type", "Data_direction"]).sum().reset_index()
            logging.debug(f"SUM Aggregated DataFrame:\n{temp_df}")
        aggregated_dataframes.append(temp_df)

    aggregated_df = pd.concat(aggregated_dataframes)
    # Pivot table to get the desired format
    logging.debug(f"Aggregated DataFrame before pivot:\n{aggregated_df.head()}\n")
    duplicates = aggregated_df[
        aggregated_df.duplicated(subset=["Row_num", "Job_name", "Job_type", "Data_direction"], keep=False)]
    logging.debug(f"Duplicate entries:\n{duplicates}")
    logging.debug(f"Unique values in 'Job_name':\n{aggregated_df['Job_name'].unique()}\t\t")
    logging.debug(f"Unique values in 'Job_type':{aggregated_df['Job_type'].unique()}\t\t")
    logging.debug(f"Unique values in 'Data_direction':\n{aggregated_df['Data_direction'].unique()}")
    result_df = aggregated_df.pivot_table(index="Row_num", columns=["Job_name", "Job_type", "Data_direction"],
                                          values="Value", aggfunc="first")
    result_df = result_df.reset_index()
    logging.debug(f"Missing values in result DataFrame:\n{result_df.isnull().sum()}")
    result_df.fillna(0, inplace=True)
    logging.debug(f"Final Result DataFrame:\n{result_df.head()}")
    logging.info(f"Finished processing client/server log files in directory: {directory}")
    return result_df


def detect_mode(directory):
    """Detect the mode based on filenames in the directory."""
    for filename in os.listdir(directory):
        if ".log." in filename:
            return "client/server"
    return "standalone"


def main():
    logging.info("Main function started.")

    # Argument parsing
    parser = argparse.ArgumentParser(description="Process FIO log files and export to a specified format.")
    parser.add_argument("directory", type=str, help="(Required) Directory where the log files are located.")
    parser.add_argument("-f", "--format", choices=["excel", "csv"], default="excel",
                        help="Export format default: excel.")
    args = parser.parse_args()

    directory = args.directory
    export_format = args.format

    logging.info(f"Processing files from directory: {directory}")

    mode = detect_mode(directory)
    logging.info(f"Detected mode: {mode}")

    if mode == "standalone":
        data_df = process_standalone_log_files(directory)
    else:
        data_df = process_client_server_log_files(directory)

    # Save data based on the provided format
    if export_format == "excel":
        filename = f"fio-results-{datetime.now().strftime('%Y-%m-%d-%H-%M')}.xlsx"
        write_to_excel(data_df, filename)
    else:
        filename = f"fio-results-{datetime.now().strftime('%Y-%m-%d-%H-%M')}.csv"
        write_to_csv(data_df, filename)

    logging.info(f"Data has been successfully exported to {filename}")
    print(f"Data has been successfully exported to {filename}")


if __name__ == "__main__":
    main()
    logging.info("Script finished.")
